import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load existing products database
json_path = os.path.join("Official Website", "assets", "data", "products.json")
products = []
if os.path.exists(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        products = json.load(f)

# Helper function to detect Silo from product name
def detect_silo(name, brand):
    name_upper = name.upper()
    silos = [
        "MERCURIAL", "SUPERFLY", "VAPOR", "PHANTOM", "TIEMPO",
        "PREDATOR", "F50", "COPA", "CRAZYFAST",
        "FUTURE", "ULTRA", "KING",
        "MORELIA", "ALPHA", "MONARCIDA",
        "TOP FLEX", "TACTICO", "DRIBLING",
        "TEKELA", "FURON", "442"
    ]
    for s in silos:
        if s in name_upper:
            return s.title()
    return "Pro Line"

# Split products into Futsal, Society, and Campo
futsal_products = []
society_products = []
campo_products = []

for p in products:
    surface = p.get("surface", "").upper()
    name = p.get("name", "").lower()
    if surface == "IC" or "futsal" in name or "indoor" in name or "ic" in name.split():
        futsal_products.append(p)
    elif surface == "TF" or "turf" in name or "society" in name or "tf" in name.split():
        society_products.append(p)
    else:
        campo_products.append(p)

print(f"Products loaded -> Futsal: {len(futsal_products)}, Society: {len(society_products)}, Campo: {len(campo_products)}")

def create_excel_catalog(filename, title, product_list, prefix, surface_code):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # Styles
    header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFD700")
    data_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    # The exact 18 columns requested by the user
    headers = [
        "sku",
        "name",
        "brand",
        "silo",
        "level",
        "category",
        "surface",
        "price",
        "color",
        "sizes",
        "description",
        "badge",
        "angle1 (main)",
        "angle2",
        "angle3",
        "angle4",
        "angle5",
        "angle6"
    ]

    # Row 1: Headers
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Populate Data
    for idx, p in enumerate(product_list, start=1):
        sku = f"GK-{prefix}-{idx:03d}"
        p["sku"] = sku  # store back in dict
        
        name = p.get("name", "")
        brand = p.get("brand", name.split()[0] if name else "Golden Kicks")
        silo = p.get("silo") or detect_silo(name, brand)
        level = p.get("level", "Elite")
        category = "Chuteiras"
        surf = p.get("surface", surface_code)
        price = p.get("price", 0)
        color = p.get("color", "White / Gold")
        sizes_str = ", ".join(str(s) for s in p.get("sizes", [35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46]))
        raw_desc = p.get("description")
        if isinstance(raw_desc, dict):
            desc = raw_desc.get("pt", "")
        else:
            desc = str(raw_desc or f"Chuteira oficial {brand} {name} de alta precisão e máximo desempenho.")
        badge = p.get("badge", "Novo")

        gallery = p.get("gallery", [])
        if not gallery and p.get("image"):
            gallery = [p.get("image")]

        angle1 = gallery[0] if len(gallery) > 0 else ""
        angle2 = gallery[1] if len(gallery) > 1 else ""
        angle3 = gallery[2] if len(gallery) > 2 else ""
        angle4 = gallery[3] if len(gallery) > 3 else ""
        angle5 = gallery[4] if len(gallery) > 4 else ""
        angle6 = gallery[5] if len(gallery) > 5 else (gallery[6] if len(gallery) > 6 else "")

        row_data = [
            sku,
            name,
            brand,
            silo,
            level,
            category,
            surf,
            price,
            color,
            sizes_str,
            desc,
            badge,
            angle1,
            angle2,
            angle3,
            angle4,
            angle5,
            angle6
        ]
        ws.append(row_data)

    # Style Data Rows
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 42
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            if col in [1, 3, 4, 5, 6, 7, 8, 12]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column Widths
    col_widths = [
        16,  # sku
        36,  # name
        14,  # brand
        16,  # silo
        12,  # level
        14,  # category
        10,  # surface
        12,  # price
        18,  # color
        30,  # sizes
        40,  # description
        12,  # badge
        35,  # angle1
        35,  # angle2
        35,  # angle3
        35,  # angle4
        35,  # angle5
        35   # angle6
    ]
    for col_i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    wb.save(filename)
    print(f"Created {filename} with {len(product_list)} products.")

create_excel_catalog("catalog_futsal.xlsx", "Futsal", futsal_products, "FUT", "IC")
create_excel_catalog("catalog_campo.xlsx", "Campo", campo_products, "CMP", "FG")
create_excel_catalog("catalog_society.xlsx", "Society", society_products, "SOC", "TF")

print("All 3 Excel catalogs updated with 18 custom columns!")
